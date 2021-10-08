import abc
import os
import pickle

import openpyxl as xls
import pandas as pd
from IPython.display import display


class BaseMapping(abc.ABC):
    fmt = {'exact_match_rate': '{:,.2%}',
           'top_n_match_rate': '{:,.2%}',
           'nobs': '{:,.0f}'}

    @classmethod
    def load_model(cls, model_file):
        """Loads the mapping instance from file."""
        with open(model_file, 'rb') as f:
            return pickle.load(f)

    def _save_model(self, check=True):
        """Dumps the object instance to self.model_file."""
        if os.path.exists(self.model_file) and check:
            with open(self.model_file, 'rb') as f:
                old = pickle.load(f)
            if len(old.mapping_candidates) > len(self.mapping_candidates):
                raise ValueError("Trying to rewrite saved model with less calculated data.")

        with open(self.model_file, 'wb') as f:
            pickle.dump(self, f)

    def _init_sheet_for_append(self, file, sheet):
        """Initialize sheet before appending it to selected XLS file."""
        if not os.path.exists(file):
            print(f'{file} file does not exist. It will be created first.\n')
            wb = xls.Workbook()
            wb.save(file)
            return

        wb = xls.load_workbook(file)
        if sheet in wb.sheetnames:
            print(f'{sheet} sheet exists in {file}. It will be deleted first.\n')
            wb.remove(wb[sheet])
            wb.save(file)

    def _save_mapping_to_xls(self, algo, sheet_name):
        """Export mapping candidates to excel file in the `self.output_data_dir` directory.

        This method will save found mapping between data fields and business terms into xls
        file in the `self.output_data_dir` directory.

        Parameters
        ----------
        algo : str
            Algorithm identifier
        sheet_name : str
            Sheet name identifier into which the mapping will be stored
        """
        excel_file = os.path.join(self.output_data_dir, 'mapping_candidates_' + self.name + '.xlsx')
        model = self.mapping_candidates[algo]
        valid = self.input.validation_mappings

        self._init_sheet_for_append(excel_file, sheet_name)

        col_names = ['business_term_' + str(i) for i in range(1, self.ncandidates + 1)]
        data_fields = pd.Series(data=model.keys(), name='data_field')
        candidates = pd.DataFrame(data=model.values(), columns=col_names)
        result = pd.concat([data_fields, candidates], axis=1)
        result[['exact_match', 'top_n_match', 'is_validation']] = 0

        if valid:
            for data_field in valid:
                candidates = [x[0] for x in model[data_field]]
                result.loc[result.data_field == data_field, 'is_validation'] = 1
                result.loc[result.data_field == data_field, 'ground_truth'] = str(valid[data_field])
                if model[data_field]:
                    if model[data_field][0][0] in valid[data_field]:
                        result.loc[result.data_field == data_field, 'exact_match'] = 1
                    if valid[data_field].intersection(candidates):
                        result.loc[result.data_field == data_field, 'top_n_match'] = \
                            min((candidates.index(i) + 1 if i in candidates else 999 for i in valid
                                [data_field]))

        with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a') as writer:
            result.to_excel(writer, sheet_name=sheet_name, index=False)

    def print_validation_results(self):
        """Generate validation results dataframe and print it out."""
        total = len(self.input.validation_mappings)
        valid = self.input.validation_mappings
        result = []

        for algo in self.mapping_candidates:
            model = self.mapping_candidates[algo]
            exact, topN = 0, 0
            for data_field in valid:
                if model[data_field]:
                    if model[data_field][0][0] in valid[data_field]:
                        exact += 1
                    if valid[data_field].intersection([x[0] for x in model[data_field]]):
                        topN += 1

            exact_rate = exact / total
            topN_rate = topN / total
            ngrams, algo_short, stemmer = algo[-8:-7], algo[:-9], algo[-1:]
            row = (algo_short, ngrams, stemmer, exact_rate, topN_rate, total, self.ncandidates)
            result.append(row)

        cols = ['algorithm', 'ngrams', 'stemmer', 'exact_match_rate', 'top_n_match_rate',
                'data_fields', 'ncandidates']
        self.validation_results = (
            pd.DataFrame(result, columns=cols).sort_values('exact_match_rate', ascending=False)
        )

        display(self.validation_results.style.hide_index().format(self.fmt))

    @abc.abstractmethod
    def load_data(self):
        return

    @abc.abstractmethod
    def find_mapping(self):
        return
